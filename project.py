import openpyxl
import mysql.connector as sql

# mysql-connector initialization
# myconn = sql.connect(host="localhost", user="root", passwd="Perambur2", database="cs_project")
# mycursor = myconn.cursor()

# openpyxl initialization
# wb = openpyxl.load_workbook("Book1.xlsx")
# ws = wb.active

# Global variables
Host = "localhost"
User = "root"
Passwd = "Perambur2"
Database = None


# Intermediate class makes it easier and convenient to work with the intermediate 2D list
class Intermediate:
    def __init__(self, list_2D):
        self.list_2D = list_2D
        self.fields = list_2D[0]
        self.records = [x for x in list_2D if x != self.fields]

    def To_Excel(self, worksheet=None, workbook_name=None):
        if worksheet != None:
            worksheet.append(self.fields)

            for rows in self.records:
                worksheet.append(rows)
        elif worksheet == None:
            wb = openpyxl.Workbook()
            ws = wb.active

            ws.append(self.fields)

            for row in self.records:
                ws.append(row)

            wb.save(filename=f"{workbook_name}.xlsx")


    def To_New_Sql(self, table_name):
        global Database

        myconn = sql.connect(host=Host, user=User, passwd=Passwd, database=Database)
        mycursor = myconn.cursor()

        f = self.fields

        mycursor.execute(f"CREATE TABLE {table_name} ({f[0]} integer, {f[1]} varchar(25), {f[2]} integer)")
            
        for row in self.records:
            mycursor.execute(f"INSERT INTO {table_name} VALUES({row[0]}, '{row[1]}', {row[2]})")
        myconn.commit()
        myconn.close()         


# functions:=
# Excel_To_Intermediate function converts the excel (.xlsx) file to the intermediate 2D list
def Excel_To_Intermediate(worksheet):
        intermediate_list = []

        for i in worksheet.values:
            intermediate_list.append(list(i))

        return Intermediate(intermediate_list)


# Sql_To_Intermediate function converts a sql table to the intermediate 2D list
def Sql_To_Intermediate(database, table):
    fields = []
    rows = []

    Database = database

    myconn = sql.connect(host=Host, user=User, passwd=Passwd, database=Database)
    mycursor = myconn.cursor()

    mycursor.execute(f"DESC {table}")
    for i in mycursor:
        fields.append(i[0])

    mycursor.execute(f"SELECT * FROM {table}")
    for i in mycursor:
        rows.append(list(i))
    
    myconn.close()
    rows.insert(0, fields)

    return Intermediate(rows)


"""When this program starts it should offer three options for what the user can do

    1. Make a new sql table using excel?
    2. Make a excel sheet from a sql table?
    3. edit a sql table using excel?"""


def user():
    print("What would you like to do?(type 7 to exit)")
    print("")
    print("1. Make a new sql table using excel?\n2. Make a excel sheet from a sql table?\n3. Edit a sql table using excel?")
    print("")
    print("type 1, 2, or 3 to choose\n")
    
    user_input = input(">> ")

    return user_input


# Making_Sql_U_Excel function makes a SQL table using a excel file
def Making_Sql_U_Excel():
    global Database

    print("What is the name of your excel file? (type it without the file extension)\n")
    excel_name = input(">> ")

    print("Choose a database\n")
    Database = input(">> ")

    print("What would you like to be the name of the table?\n")
    table_name = input(">> ")

    wb = openpyxl.load_workbook(f"{excel_name}.xlsx")
    ws = wb.active

    Inter = Excel_To_Intermediate(ws)

    Inter.To_New_Sql(table_name)


# Making_Excel_U_Sql function makes a excel file using a SQL table
def Making_Excel_U_Sql():
    print("Choose a database\n")
    Database = input(">> ")

    print("Which table do you want to convert to excel?\n")
    table_name = input(">> ")

    print("name your excel file (type it without the file extension)\n")
    excel_name = input(">> ")

    wb = openpyxl.Workbook()
    ws = wb.active

    Inter = Sql_To_Intermediate(Database, table_name)

    Inter.To_Excel(None, excel_name)


def Edit_Sql_U_Excel():
    print("Choose a database\n")
    Database = input(">> ")

    print("Which table do you want to edit using excel?\n")
    table_name = input(">> ")

    Inter = Sql_To_Intermediate(Database, table_name)

    Inter.To_Excel(None, table_name)



def main():
    while True:
        u = user()
        if u == '1':
            Making_Sql_U_Excel()
        elif u == '2':
            Making_Excel_U_Sql()
        elif u == '3':
            Edit_Sql_U_Excel()
        elif u == '7':
            break
        else:
            print("Invalid input, try again.\n\n")


main()








